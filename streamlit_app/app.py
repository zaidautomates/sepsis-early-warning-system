# ─────────────────────────────────────────────────────────────────────────────
# Sepsis Early Warning System — ICU Dashboard v6.0
# Author : Zaid Ali | Roll No. 40 | BS CS 6th Semester | AWKUM 2025–26
# Model  : XGBoost | AUROC 0.8158 | 6-Hour Early Prediction
# ─────────────────────────────────────────────────────────────────────────────
import streamlit as st, numpy as np, joblib, json, shap, time, io, base64, os
import streamlit.components.v1 as st_html
import matplotlib; matplotlib.use('Agg')
import matplotlib.pyplot as plt, matplotlib.patches as mpatches
import pandas as pd
from datetime import datetime, timedelta
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors as rl_colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_LEFT

st.set_page_config(page_title="Sepsis EWS — Clinical Dashboard", page_icon="🏥",
                   layout="wide", initial_sidebar_state="expanded")

# ═══ GLOBAL CSS ═══
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=Space+Grotesk:wght@500;600;700&family=JetBrains+Mono:wght@400;500;600&display=swap');
:root {
  --bg: #f3f6f9; --surface: #ffffff; --surface2: #f8fafc; --border: #e2e8f0; --border2: #cbd5e1;
  --text: #0f172a; --text2: #475569; --text3: #64748b;
  --teal: #0ea5e9; --teal-lt: #e0f2fe; --teal-glow: rgba(14,165,233,0.3);
  --red: #ef4444; --red-lt: #fee2e2; --red-glow: rgba(239,68,68,0.3);
  --amber: #f59e0b; --amber-lt: #fef3c7; --amber-glow: rgba(245,158,11,0.3);
  --green: #10b981; --green-lt: #d1fae5; --blue: #3b82f6; --blue-lt: #dbeafe;
  --shadow-sm: 0 2px 8px -2px rgba(15,23,42,0.06), 0 4px 12px -4px rgba(15,23,42,0.04);
  --shadow-md: 0 8px 24px -6px rgba(15,23,42,0.08), 0 12px 32px -8px rgba(15,23,42,0.06);
  --shadow-lg: 0 16px 40px -10px rgba(15,23,42,0.1), 0 24px 64px -16px rgba(15,23,42,0.08);
  --radius: 16px; --radius-sm: 10px;
}
@media (prefers-color-scheme: dark) {
  :root {
    --bg: #0b0f19; --surface: #111827; --surface2: #1f2937; --border: #374151; --border2: #4b5563;
    --text: #f8fafc; --text2: #cbd5e1; --text3: #94a3b8;
    --teal: #38bdf8; --teal-lt: rgba(56,189,248,0.1);
    --red: #f87171; --red-lt: rgba(248,113,113,0.1);
    --amber: #fbbf24; --amber-lt: rgba(251,191,36,0.1);
    --green: #34d399; --green-lt: rgba(52,211,153,0.1); --blue: #60a5fa; --blue-lt: rgba(96,165,250,0.1);
    --shadow-sm: 0 2px 10px rgba(0,0,0,0.3); --shadow-md: 0 8px 30px rgba(0,0,0,0.4); --shadow-lg: 0 16px 50px rgba(0,0,0,0.5);
  }
}
html, body, [class*="css"] { font-family: 'Inter', sans-serif; color: var(--text); -webkit-font-smoothing: antialiased; }
.stApp { background: var(--bg); }

/* Improved Accessibility Focus States */
a:focus, button:focus, input:focus, [tabindex]:focus { outline: 2px solid var(--teal) !important; outline-offset: 2px; }

/* Sidebar Styling */
section[data-testid="stSidebar"] { background: rgba(var(--surface), 0.6) !important; backdrop-filter: blur(20px); -webkit-backdrop-filter: blur(20px); border-right: 1px solid var(--border) !important; }
section[data-testid="stSidebar"] label { color: var(--text2) !important; font-size: 11px !important; font-weight: 600 !important; letter-spacing: 0.5px !important; text-transform: uppercase; }
section[data-testid="stSidebar"] input[type="number"] { background: var(--surface2) !important; border: 1px solid var(--border) !important; color: var(--text) !important; border-radius: 8px !important; font-family: 'JetBrains Mono', monospace !important; font-size: 14px !important; transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1); box-shadow: inset 0 1px 3px rgba(0,0,0,0.02); }
section[data-testid="stSidebar"] input[type="number"]:focus { border-color: var(--teal) !important; background: var(--surface) !important; box-shadow: 0 0 0 4px var(--teal-lt), inset 0 1px 3px rgba(0,0,0,0.02) !important; transform: translateY(-1px); }

/* Advanced Animations */
@keyframes fadeUp { from { opacity: 0; transform: translateY(16px) } to { opacity: 1; transform: translateY(0) } }
@keyframes fadeScale { from { opacity: 0; transform: scale(0.96) } to { opacity: 1; transform: scale(1) } }
@keyframes pulseRed { 0%, 100% { box-shadow: 0 0 0 0 rgba(239,68,68,0.4) } 70% { box-shadow: 0 0 0 16px rgba(239,68,68,0) } }
@keyframes pulseAmber { 0%, 100% { box-shadow: 0 0 0 0 rgba(245,158,11,0.3) } 70% { box-shadow: 0 0 0 14px rgba(245,158,11,0) } }
@keyframes shimmer { 0% { left: -100% } 100% { left: 200% } }
@keyframes float { 0%, 100% { transform: translateY(0) } 50% { transform: translateY(-8px) } }
@keyframes glowSweep { 0% { background-position: 0% 50%; } 50% { background-position: 100% 50%; } 100% { background-position: 0% 50%; } }

/* Header Branding (Glass & Gradient) */
.hdr-wrap { background: linear-gradient(135deg, #0f172a 0%, #1e3a8a 50%, #0891b2 100%); background-size: 200% 200%; border-radius: var(--radius); padding: 36px 40px; color: #fff; position: relative; overflow: hidden; margin-bottom: 20px; box-shadow: var(--shadow-md); animation: fadeScale .5s ease-out, glowSweep 15s ease infinite; border: 1px solid rgba(255,255,255,0.1); }
.hdr-wrap::before { content: ''; position: absolute; top: -50%; left: -20%; width: 300px; height: 300px; border-radius: 50%; background: radial-gradient(circle, rgba(255,255,255,0.15) 0%, transparent 70%); animation: float 10s ease-in-out infinite; filter: blur(20px); }
.hdr-wrap::after { content: ''; position: absolute; bottom: -30%; right: -10%; width: 250px; height: 250px; border-radius: 50%; background: radial-gradient(circle, rgba(14,165,233,0.3) 0%, transparent 70%); animation: float 8s ease-in-out infinite reverse; filter: blur(15px); }
.hdr-title { font-family: 'Space Grotesk', sans-serif; font-size: 40px; font-weight: 700; color: #fff; margin: 0; letter-spacing: -0.5px; line-height: 1.1; position: relative; z-index: 2; text-shadow: 0 2px 10px rgba(0,0,0,0.3); }
.hdr-sub { font-size: 15px; color: rgba(255,255,255,.85); margin: 8px 0 0; line-height: 1.6; position: relative; z-index: 2; font-weight: 300; }
.hdr-badge { display: inline-flex; align-items: center; background: rgba(0,0,0,0.25); border: 1px solid rgba(255,255,255,.15); backdrop-filter: blur(8px); -webkit-backdrop-filter: blur(8px); border-radius: 24px; padding: 6px 14px; font-size: 11.5px; font-weight: 600; color: #fff; margin-right: 12px; margin-top: 16px; text-transform: uppercase; letter-spacing: 1px; font-family: 'JetBrains Mono', monospace; position: relative; z-index: 2; transition: transform 0.3s; }
.hdr-badge:hover { transform: translateY(-2px); background: rgba(0,0,0,0.4); }
.hdr-live { background: rgba(239,68,68,0.2); border-color: rgba(239,68,68,0.4); color: #fecaca; }
.hdr-live::before { content: ''; display: inline-block; width: 6px; height: 6px; background: #ef4444; border-radius: 50%; margin-right: 8px; box-shadow: 0 0 8px #ef4444; animation: blink 2s infinite; }

/* Cards & Layout */
.stat-card { background: var(--surface); border: 1px solid var(--border); border-top: 4px solid var(--teal); border-radius: var(--radius); padding: 22px 18px; text-align: center; box-shadow: var(--shadow-sm); transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1); animation: fadeUp .5s ease-out both; position: relative; overflow: hidden; }
.stat-card:hover { transform: translateY(-5px); box-shadow: var(--shadow-lg); border-color: var(--teal); }
.stat-card::after { content:''; position:absolute; top:0; left:0; width:100%; height:100%; background: linear-gradient(180deg, rgba(255,255,255,0.05) 0%, transparent 100%); pointer-events:none; }
.stat-num { font-size: 26px; font-weight: 700; color: var(--text); margin: 0; font-family: 'Space Grotesk', sans-serif; background: linear-gradient(135deg, var(--teal), #3b82f6); -webkit-background-clip: text; -webkit-text-fill-color: transparent; }
.stat-lbl { font-size: 11px; color: var(--text3); margin: 6px 0 0; text-transform: uppercase; letter-spacing: 1.2px; font-weight: 600; }

/* Risk Card Focus */
.risk-card { border-radius: var(--radius); padding: 36px 24px; text-align: center; border: 2px solid; box-shadow: var(--shadow-md); transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1); animation: fadeScale .5s ease-out; position: relative; overflow: hidden; }
.risk-card:hover { transform: translateY(-6px); box-shadow: var(--shadow-lg); }
.risk-pct { font-size: 88px; font-weight: 700; margin: 0; line-height: 1; letter-spacing: -4px; font-family: 'Space Grotesk', sans-serif; position: relative; z-index: 2; text-shadow: 0 4px 16px rgba(0,0,0,0.1); }
.risk-level { font-size: 14px; font-weight: 700; margin: 16px 0 8px; text-transform: uppercase; letter-spacing: 4px; position: relative; z-index: 2; }
.risk-msg { font-size: 14px; margin: 4px 0 0; line-height: 1.6; font-weight: 500; position: relative; z-index: 2; opacity: 0.9; }

.risk-high { background: linear-gradient(135deg, var(--red-lt), var(--surface)); border-color: var(--red); color: #7f1d1d; animation: pulseRed 2.5s infinite, fadeScale .5s ease-out; }
@media(prefers-color-scheme: dark) { .risk-high { background: linear-gradient(135deg, rgba(239,68,68,0.15), var(--surface)); color: #fca5a5; } }
.risk-high .risk-pct { color: var(--red); }
.risk-high::before { content:''; position:absolute; top:-50%; left:-50%; width:200%; height:200%; background: radial-gradient(circle, var(--red-glow) 0%, transparent 50%); opacity:0.5; pointer-events:none; z-index:1; }

.risk-mod { background: linear-gradient(135deg, var(--amber-lt), var(--surface)); border-color: var(--amber); color: #78350f; animation: pulseAmber 2.8s infinite, fadeScale .5s ease-out; }
@media(prefers-color-scheme: dark) { .risk-mod { background: linear-gradient(135deg, rgba(245,158,11,0.15), var(--surface)); color: #fcd34d; } }
.risk-mod .risk-pct { color: var(--amber); }
.risk-mod::before { content:''; position:absolute; top:-50%; left:-50%; width:200%; height:200%; background: radial-gradient(circle, var(--amber-glow) 0%, transparent 50%); opacity:0.4; pointer-events:none; z-index:1; }

.risk-low { background: linear-gradient(135deg, var(--green-lt), var(--surface)); border-color: var(--green); color: #064e3b; box-shadow: var(--shadow-sm); }
@media(prefers-color-scheme: dark) { .risk-low { background: linear-gradient(135deg, rgba(16,185,129,0.15), var(--surface)); color: #6ee7b7; } }
.risk-low .risk-pct { color: var(--green); }

/* Progress Track Enhanced */
.prog-track { background: var(--surface2); border: 1px solid var(--border); border-radius: 12px; height: 14px; overflow: hidden; margin: 18px 0 8px; box-shadow: inset 0 2px 4px rgba(0,0,0,.04); position: relative; }
.prog-fill { height: 100%; border-radius: 12px; position: relative; overflow: hidden; transition: width 1.2s cubic-bezier(.4,0,.2,1); }
.prog-fill::after { content: ''; position: absolute; top: 0; left: 0; width: 60%; height: 100%; background: linear-gradient(90deg, transparent, rgba(255,255,255,.5), transparent); animation: shimmer 2s cubic-bezier(0.4, 0, 0.2, 1) infinite; }
.prog-labels { display: flex; justify-content: space-between; font-size: 11px; color: var(--text3); font-family: 'JetBrains Mono', monospace; font-weight: 600; padding: 0 2px; }

/* Subheaders */
.s-hdr { font-family: 'Space Grotesk', sans-serif; font-size: 12px; font-weight: 700; color: var(--teal); text-transform: uppercase; letter-spacing: 2.5px; margin: 0 0 20px; padding-bottom: 12px; border-bottom: 2px solid var(--border); position: relative; }
.s-hdr::after { content: ''; position: absolute; left: 0; bottom: -2px; width: 60px; height: 2px; background: var(--teal); border-radius: 2px; }

/* Score Boxes */
.score-box { background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius); padding: 26px 18px; text-align: center; box-shadow: var(--shadow-sm); transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1); position: relative; overflow: hidden; animation: fadeUp .6s ease-out both; }
.score-box:hover { transform: translateY(-6px) scale(1.02); box-shadow: var(--shadow-lg); border-color: transparent; }
.score-box::before { content:''; position:absolute; top:0; left:0; width:100%; height:100%; background: linear-gradient(180deg, rgba(255,255,255,0.05) 0%, transparent 100%); pointer-events:none; }
.score-num { font-size: 38px; font-weight: 800; margin: 0; font-family: 'Space Grotesk', sans-serif; line-height: 1; text-shadow: 0 2px 8px rgba(0,0,0,0.05); }
.score-lbl { font-size: 12px; font-weight: 700; margin: 12px 0 0; text-transform: uppercase; letter-spacing: 1.5px; color: var(--text3); }

/* Premium Vital Chips (Glassmorphism) */
.vital-chip { border-radius: var(--radius); padding: 18px 14px; text-align: center; border: 1px solid rgba(255,255,255,0.4); background: rgba(255,255,255,0.7); backdrop-filter: blur(16px); -webkit-backdrop-filter: blur(16px); animation: fadeUp .4s ease-out both; transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1); box-shadow: var(--shadow-sm); cursor: default; }
@media (prefers-color-scheme: dark) { .vital-chip { background: rgba(30,41,59,0.5); border: 1px solid rgba(255,255,255,0.05); } }
.vital-chip:hover { transform: translateY(-4px) scale(1.02); box-shadow: var(--shadow-lg); background: var(--surface); border-color: var(--teal); z-index: 10; }
.vc-normal { border-left: 4px solid var(--green); }
.vc-warn { border-left: 4px solid var(--amber); }
.vc-critical { border-left: 4px solid var(--red); }

.vital-name { font-size: 10.5px; font-weight: 700; color: var(--text3); margin: 0; text-transform: uppercase; letter-spacing: 1px; font-family: 'Inter', sans-serif; }
.vital-val { font-size: 22px; font-weight: 700; margin: 6px 0 2px; font-family: 'Space Grotesk', sans-serif; color: var(--text); }
.vital-unit { font-size: 11px; color: var(--text2); font-weight: 500; }
.vital-range { font-size: 10px; color: var(--text3); margin: 6px 0 0; font-family: 'JetBrains Mono', monospace; opacity: 0.8; background: var(--surface2); padding: 4px 8px; border-radius: 6px; display: inline-block; }

/* Miscellaneous Components */
.sirs-wrap { background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius); padding: 20px 24px; box-shadow: var(--shadow-sm); transition: box-shadow 0.3s; }
.sirs-wrap:hover { box-shadow: var(--shadow-md); }
.sirs-row { display: flex; align-items: center; gap: 14px; padding: 10px 0; border-bottom: 1px solid var(--border); font-size: 13.5px; font-weight: 500; color: var(--text); transition: background 0.2s; border-radius: 8px; }
.sirs-row:hover { background: var(--surface2); padding-left: 8px; padding-right: 8px; margin: 0 -8px; }
.sirs-row:last-child { border-bottom: none; }
.sirs-dot { width: 12px; height: 12px; border-radius: 50%; flex-shrink: 0; box-shadow: inset 0 2px 4px rgba(0,0,0,0.1); }

/* Recommendation Cards */
.rec-card { background: var(--surface); border-left: 4px solid; border-radius: var(--radius-sm); padding: 16px 20px; margin: 10px 0; font-size: 14px; line-height: 1.6; animation: fadeUp .4s ease-out both; box-shadow: var(--shadow-sm); transition: all 0.3s; position: relative; overflow: hidden; }
.rec-card:hover { transform: translateX(6px); box-shadow: var(--shadow-md); }
.rec-title { font-size: 11px; font-weight: 700; text-transform: uppercase; letter-spacing: 1.5px; margin-bottom: 6px; font-family: 'Inter', sans-serif; }
.rc-high { border-color: var(--red); color: #7f1d1d; background: linear-gradient(90deg, var(--red-lt), var(--surface)); }
.rc-mod { border-color: var(--amber); color: #78350f; background: linear-gradient(90deg, var(--amber-lt), var(--surface)); }
.rc-low { border-color: var(--green); color: #064e3b; background: linear-gradient(90deg, var(--green-lt), var(--surface)); }
.rc-info { border-color: var(--teal); color: #164e63; background: linear-gradient(90deg, var(--blue-lt), var(--surface)); }
@media (prefers-color-scheme: dark) {
  .rc-high{background: linear-gradient(90deg, rgba(220,38,38,0.15), var(--surface)); color: #fca5a5; }
  .rc-mod{background: linear-gradient(90deg, rgba(217,119,6,0.15), var(--surface)); color: #fcd34d; }
  .rc-low{background: linear-gradient(90deg, rgba(5,150,105,0.15), var(--surface)); color: #6ee7b7; }
  .rc-info{background: linear-gradient(90deg, rgba(8,145,178,0.15), var(--surface)); color: #67e8f9; }
}

/* Sidebar Advanced Styling */
.sidebar-header { text-align: center; padding: 24px 0; background: linear-gradient(180deg, var(--teal-lt) 0%, transparent 100%); border-radius: var(--radius); margin-bottom: 20px; border: 1px solid rgba(14,165,233,0.1); box-shadow: inset 0 2px 10px rgba(255,255,255,0.5); }
@media (prefers-color-scheme: dark) { .sidebar-header { box-shadow: inset 0 2px 10px rgba(255,255,255,0.02); } }
.sidebar-icon { font-size: 48px; filter: drop-shadow(0 4px 8px rgba(0,0,0,0.15)); animation: float 6s ease-in-out infinite; }
.sidebar-title { font-size: 16px; font-weight: 800; color: var(--text); margin-top: 12px; letter-spacing: 0.5px; font-family: 'Space Grotesk', sans-serif; }
.sidebar-sub { font-size: 10px; color: var(--teal); margin-top: 4px; font-family: 'JetBrains Mono', monospace; font-weight: 700; letter-spacing: 1.5px; text-transform: uppercase; opacity: 0.8; }

/* Organ Risk Cards */
.organ-risk-card { background: var(--surface); border: 2px solid; border-radius: var(--radius); padding: 28px 24px; text-align: center; margin-top: 10px; box-shadow: var(--shadow-sm); transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1); position: relative; overflow: hidden; animation: fadeScale .5s ease-out; }
.organ-risk-card:hover { transform: translateY(-6px) scale(1.02); box-shadow: var(--shadow-lg); }
.organ-risk-title { font-size: 13px; font-weight: 800; color: var(--text2); text-transform: uppercase; letter-spacing: 2px; }
.organ-risk-value { font-size: 56px; font-weight: 900; font-family: 'Space Grotesk', sans-serif; margin: 8px 0; text-shadow: 0 4px 12px rgba(0,0,0,0.1); line-height: 1; }
.organ-risk-badge { display: inline-block; font-size: 11px; font-weight: 700; padding: 6px 16px; border-radius: 24px; text-transform: uppercase; letter-spacing: 1.5px; border: 1px solid; box-shadow: inset 0 1px 3px rgba(255,255,255,0.2); }

/* Biomarker Rows */
.biomarker-row { display: flex; justify-content: space-between; align-items: center; padding: 12px 16px; border-left: 4px solid; background: var(--surface2); border-radius: 0 10px 10px 0; margin: 8px 0; transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1); box-shadow: 0 1px 2px rgba(0,0,0,0.02); border-top: 1px solid transparent; border-right: 1px solid transparent; border-bottom: 1px solid transparent; }
.biomarker-row:hover { background: var(--surface); transform: translateX(6px); box-shadow: var(--shadow-md); border-color: var(--border); }
.biomarker-label { font-size: 13px; font-weight: 700; color: var(--text); }
.biomarker-val { font-family: 'JetBrains Mono', monospace; font-weight: 800; font-size: 14px; }

/* Alert Strips */
.alert-strip { border-radius: var(--radius-sm); padding: 14px 20px; font-size: 13px; font-weight: 700; text-align: center; text-transform: uppercase; letter-spacing: 1.5px; margin: 12px 0; font-family: 'Inter', sans-serif; border: 1px solid; box-shadow: var(--shadow-sm); }
.a-high { background: var(--red-lt); border-color: rgba(239,68,68,0.3); color: #7f1d1d; }
.a-mod { background: var(--amber-lt); border-color: rgba(245,158,11,0.3); color: #78350f; }
.a-low { background: var(--green-lt); border-color: rgba(16,185,129,0.3); color: #064e3b; }
.a-info { background: var(--blue-lt); border-color: rgba(59,130,246,0.3); color: #1e3a8a; }

/* Info Boxes & Steps */
.i-box { background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius); padding: 20px 24px; font-size: 14px; color: var(--text2); line-height: 1.7; box-shadow: var(--shadow-sm); transition: box-shadow 0.3s; }
.i-box:hover { box-shadow: var(--shadow-md); }
.scores-box { background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius); padding: 20px 24px; box-shadow: var(--shadow-sm); transition: box-shadow 0.3s; }
.scores-box:hover { box-shadow: var(--shadow-md); }
.score-row { display: flex; justify-content: space-between; align-items: center; padding: 10px 0; border-bottom: 1px dashed var(--border); font-size: 14px; } .score-row:last-child { border-bottom: none; }
.score-label { color: var(--text2); font-weight: 500; } .score-value { font-family: 'JetBrains Mono', monospace; font-weight: 700; font-size: 15px; color: var(--text); background: var(--surface2); padding: 2px 8px; border-radius: 6px; }

.step-card { background: var(--surface); border: 1px solid var(--border); border-radius: var(--radius); padding: 28px 24px; text-align: center; transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1); box-shadow: var(--shadow-sm); }
.step-card:hover { border-color: var(--teal); box-shadow: var(--shadow-lg); transform: translateY(-5px); }
.step-n { font-size: 42px; font-weight: 700; color: var(--teal); margin: 0; font-family: 'Space Grotesk', sans-serif; background: linear-gradient(135deg, var(--teal), var(--blue)); -webkit-background-clip: text; -webkit-text-fill-color: transparent; }
.step-t { font-size: 15px; font-weight: 700; color: var(--text); margin: 12px 0 8px; font-family: 'Inter', sans-serif; }
.step-d { font-size: 13px; color: var(--text2); line-height: 1.6; }

.summary-bar { border-radius: var(--radius-sm); padding: 22px 28px; font-size: 15px; font-weight: 600; text-align: center; letter-spacing: 0.5px; font-family: 'Inter', sans-serif; margin-top: 20px; animation: fadeUp .6s ease-out; border: 1px solid; box-shadow: var(--shadow-md); }

/* Tab Overrides (Sleek Segmented Control Look) */
.stTabs [data-baseweb="tab-list"] { background: var(--surface2); border-radius: 12px; padding: 6px; border: 1px solid var(--border); gap: 8px; box-shadow: inset 0 2px 4px rgba(0,0,0,0.02); }
.stTabs [data-baseweb="tab"] { border-radius: 8px; font-weight: 600; font-size: 14px; color: var(--text2); padding: 12px 20px; transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1); border: 1px solid transparent; }
.stTabs [data-baseweb="tab"]:hover { color: var(--text); background: rgba(255,255,255,0.5); }
@media (prefers-color-scheme: dark) { .stTabs [data-baseweb="tab"]:hover { background: rgba(255,255,255,0.05); } }
.stTabs [aria-selected="true"] { background: var(--surface) !important; color: var(--teal) !important; box-shadow: var(--shadow-sm); border: 1px solid var(--border); }

/* DataFrames */
[data-testid="stDataFrame"] { border-radius: var(--radius); overflow: hidden; border: 1px solid var(--border); box-shadow: var(--shadow-sm); transition: box-shadow 0.3s; }
[data-testid="stDataFrame"]:hover { box-shadow: var(--shadow-md); }

.footer { text-align: center; font-size: 12px; color: var(--text3); margin-top: 48px; padding-top: 24px; border-top: 1px solid var(--border); font-family: 'JetBrains Mono', monospace; line-height: 1.6; opacity: 0.8; transition: opacity 0.3s; }
.footer:hover { opacity: 1; }
</style>

""", unsafe_allow_html=True)

# ═══ PREMIUM UI ENHANCEMENTS ═══
st.markdown("""
<style>
/* ── New tokens ── */
:root {
  --purple: #8b5cf6; --purple-lt: rgba(139,92,246,0.12); --purple-glow: rgba(139,92,246,0.3);
  --grad-main: linear-gradient(135deg, #0ea5e9 0%, #3b82f6 50%, #8b5cf6 100%);
  --grad-main-rev: linear-gradient(135deg, #8b5cf6 0%, #3b82f6 50%, #0ea5e9 100%);
  --shadow-glow: 0 0 24px rgba(14,165,233,0.25), 0 8px 32px rgba(59,130,246,0.15);
  --shadow-purple: 0 0 24px rgba(139,92,246,0.25), 0 8px 32px rgba(139,92,246,0.15);
}

/* ── New keyframes ── */
@keyframes blink { 0%,100%{opacity:1} 50%{opacity:0.3} }
@keyframes gradFlow { 0%{background-position:0% 50%} 50%{background-position:100% 50%} 100%{background-position:0% 50%} }
@keyframes ripple { 0%{transform:scale(0);opacity:0.6} 100%{transform:scale(2.5);opacity:0} }
@keyframes progGlow { 0%,100%{box-shadow:0 0 8px rgba(14,165,233,0.6)} 50%{box-shadow:0 0 20px rgba(139,92,246,0.8)} }
@keyframes sectionIn { from{opacity:0;transform:translateY(12px) scale(0.99)} to{opacity:1;transform:none} }
@keyframes countUp { from{opacity:0;transform:translateY(8px)} to{opacity:1;transform:none} }

/* ── Streamlit button → pill gradient ── */
.stButton > button {
  border-radius: 50px !important;
  background: var(--grad-main) !important;
  background-size: 200% 200% !important;
  color: #fff !important;
  border: none !important;
  font-weight: 700 !important;
  font-family: 'Inter', sans-serif !important;
  letter-spacing: 0.3px !important;
  padding: 10px 28px !important;
  transition: all 0.35s cubic-bezier(0.4,0,0.2,1) !important;
  box-shadow: 0 4px 15px rgba(14,165,233,0.35) !important;
  position: relative !important;
  overflow: hidden !important;
  animation: gradFlow 6s ease infinite !important;
}
.stButton > button:hover {
  transform: translateY(-2px) scale(1.03) !important;
  box-shadow: var(--shadow-glow) !important;
  background-position: 100% 50% !important;
}
.stButton > button:active {
  transform: translateY(0) scale(0.97) !important;
  box-shadow: 0 2px 8px rgba(14,165,233,0.2) !important;
}
/* ripple layer */
.stButton > button::after {
  content: '' !important;
  position: absolute !important;
  width: 100% !important; height: 100% !important;
  top: 0 !important; left: 0 !important;
  background: radial-gradient(circle, rgba(255,255,255,0.4) 0%, transparent 60%) !important;
  transform: scale(0) !important;
  opacity: 0 !important;
  transition: transform 0.5s, opacity 0.5s !important;
  border-radius: 50px !important;
}
.stButton > button:active::after { animation: ripple 0.5s ease-out !important; }

/* ── Sidebar analyze button — accent ── */
section[data-testid="stSidebar"] .stButton > button {
  background: linear-gradient(135deg, #0f172a 0%, #1e3a8a 50%, #0891b2 100%) !important;
  background-size: 200% 200% !important;
  animation: gradFlow 8s ease infinite !important;
}

/* ── Tab: animated gradient active indicator ── */
.stTabs [data-baseweb="tab-list"] {
  background: rgba(248,250,252,0.8) !important;
  backdrop-filter: blur(12px) !important;
  -webkit-backdrop-filter: blur(12px) !important;
  border-radius: 14px !important;
  padding: 6px !important;
  border: 1px solid var(--border) !important;
  gap: 6px !important;
}
.stTabs [data-baseweb="tab"] {
  border-radius: 10px !important;
  font-weight: 600 !important;
  font-size: 13px !important;
  color: var(--text2) !important;
  padding: 10px 18px !important;
  transition: all 0.35s cubic-bezier(0.4,0,0.2,1) !important;
  border: 1px solid transparent !important;
  position: relative !important;
}
.stTabs [data-baseweb="tab"]:hover {
  color: var(--teal) !important;
  background: rgba(14,165,233,0.08) !important;
  transform: translateY(-1px) !important;
}
.stTabs [aria-selected="true"] {
  background: var(--grad-main) !important;
  background-size: 200% 200% !important;
  color: #fff !important;
  box-shadow: 0 4px 14px rgba(14,165,233,0.4) !important;
  border: none !important;
  animation: gradFlow 5s ease infinite !important;
}

/* ── Glassmorphism panels ── */
.stat-card, .score-box, .organ-risk-card, .vital-chip, .sirs-wrap, .scores-box, .i-box {
  background: rgba(255,255,255,0.72) !important;
  backdrop-filter: blur(18px) !important;
  -webkit-backdrop-filter: blur(18px) !important;
  border: 1px solid rgba(255,255,255,0.55) !important;
}
@media (prefers-color-scheme: dark) {
  .stat-card, .score-box, .organ-risk-card, .vital-chip, .sirs-wrap, .scores-box, .i-box {
    background: rgba(17,24,39,0.65) !important;
    border: 1px solid rgba(255,255,255,0.06) !important;
  }
}

/* ── Stat-card glow on hover ── */
.stat-card:hover {
  box-shadow: var(--shadow-glow) !important;
  border-color: rgba(14,165,233,0.4) !important;
  transform: translateY(-6px) !important;
}
.score-box:hover {
  box-shadow: var(--shadow-purple) !important;
  border-color: rgba(139,92,246,0.3) !important;
}

/* ── Progress bar glow sweep ── */
.prog-fill {
  background: var(--grad-main) !important;
  background-size: 200% 200% !important;
  animation: gradFlow 4s ease infinite, progGlow 3s ease-in-out infinite !important;
  transition: width 1.4s cubic-bezier(0.4,0,0.2,1) !important;
}

/* ── Smooth section entry ── */
[data-testid="stVerticalBlock"] > div { animation: sectionIn 0.45s ease-out both; }

/* ── Stat numbers animate in ── */
.stat-num, .score-num, .risk-pct, .organ-risk-value { animation: countUp 0.6s ease-out both; }

/* ── Summary bar gradient ── */
.a-high { background: linear-gradient(135deg, #fee2e2, #fecaca) !important; border-color: rgba(239,68,68,0.4) !important; }
.a-mod  { background: linear-gradient(135deg, #fef3c7, #fde68a) !important; border-color: rgba(245,158,11,0.4) !important; }
.a-low  { background: linear-gradient(135deg, #d1fae5, #a7f3d0) !important; border-color: rgba(16,185,129,0.4) !important; }
.a-info { background: linear-gradient(135deg, #dbeafe, #bfdbfe) !important; border-color: rgba(59,130,246,0.4) !important; }

/* ── Summary bar itself ── */
.summary-bar {
  background: var(--grad-main) !important;
  background-size: 200% 200% !important;
  animation: gradFlow 6s ease infinite, fadeUp 0.6s ease-out !important;
  color: #fff !important;
  border: none !important;
  box-shadow: var(--shadow-glow) !important;
  border-radius: 14px !important;
}
.summary-bar.a-high { background: linear-gradient(135deg,#dc2626,#b91c1c,#991b1b) !important; }
.summary-bar.a-mod  { background: linear-gradient(135deg,#d97706,#b45309,#92400e) !important; }
.summary-bar.a-low  { background: linear-gradient(135deg,#059669,#047857,#065f46) !important; }

/* ── Rec-card slide enhancement ── */
.rec-card { transition: transform 0.35s cubic-bezier(0.4,0,0.2,1), box-shadow 0.35s ease !important; }
.rec-card:hover { transform: translateX(8px) !important; box-shadow: var(--shadow-glow) !important; }

/* ── Step card gradient border on hover ── */
.step-card { transition: all 0.4s cubic-bezier(0.4,0,0.2,1) !important; }
.step-card:hover { border-color: transparent !important; box-shadow: var(--shadow-glow) !important; outline: 2px solid rgba(14,165,233,0.5) !important; }

/* ── Footer gradient line ── */
.footer { border-top: 2px solid transparent !important; border-image: var(--grad-main) 1 !important; }

/* ── Download buttons ── */
[data-testid="stDownloadButton"] > button {
  border-radius: 50px !important;
  background: linear-gradient(135deg,#0f172a,#1e3a8a,#0891b2) !important;
  background-size: 200% 200% !important;
  color: #fff !important; border: none !important;
  font-weight: 700 !important;
  transition: all 0.35s ease !important;
  box-shadow: 0 4px 14px rgba(8,145,178,0.3) !important;
  animation: gradFlow 8s ease infinite !important;
}
[data-testid="stDownloadButton"] > button:hover {
  transform: translateY(-2px) scale(1.03) !important;
  box-shadow: var(--shadow-glow) !important;
}
</style>
""", unsafe_allow_html=True)


# ═══ MODEL LOADING ═══
_APP_DIR = os.path.dirname(os.path.abspath(__file__))
_MODEL_DIR = os.path.join(_APP_DIR, '..', 'models')

@st.cache_resource
def load_models():
    m=joblib.load(os.path.join(_MODEL_DIR,'xgboost_model.pkl'))
    s=joblib.load(os.path.join(_MODEL_DIR,'scaler.pkl'))
    f=json.load(open(os.path.join(_MODEL_DIR,'feature_list.json')))
    e=shap.TreeExplainer(m)
    return m,s,f,e
try:
    model,scaler,features,explainer=load_models()
except Exception as e:
    st.error(f"Model load failed: {e}"); st.stop()

# ═══ CLINICAL DATA ═══
CLINICAL_RANGES={
    'HR':(60,100,20,220,'bpm'),'O2Sat':(95,100,50,100,'%'),'Temp':(36.1,37.2,32,42,'°C'),
    'SBP':(90,130,40,250,'mmHg'),'MAP':(70,105,20,180,'mmHg'),'DBP':(60,85,20,150,'mmHg'),
    'Resp':(12,20,4,60,'/min'),'WBC':(4.5,11,0.5,100,'K/µL'),'Creatinine':(0.7,1.2,0.1,20,'mg/dL'),
    'Bilirubin_total':(0.2,1.2,0.1,30,'mg/dL'),'Lactate':(0.5,2,0.1,25,'mmol/L'),
    'Glucose':(70,100,10,800,'mg/dL'),'Hgb':(12,17.5,3,25,'g/dL'),'pH':(7.35,7.45,6.8,7.8,''),
    'Age':(18,90,18,110,'yrs'),'HospAdmTime':(-200,0,-600,0,'hr'),'ICULOS':(0,10,0,90,'days'),
}
FEATURE_LABELS={'HR':'Heart Rate','O2Sat':'O₂ Saturation','Temp':'Temperature','SBP':'Systolic BP',
    'MAP':'Mean Art. Pr.','DBP':'Diastolic BP','Resp':'Resp. Rate','WBC':'WBC Count',
    'Creatinine':'Creatinine','Bilirubin_total':'Bilirubin','Lactate':'Lactate','Glucose':'Glucose',
    'Hgb':'Hemoglobin','pH':'Blood pH','Age':'Patient Age','HospAdmTime':'Adm. Offset','ICULOS':'ICU LOS'}
DEFAULTS={'HR':85,'O2Sat':97,'Temp':37,'SBP':120,'MAP':85,'DBP':70,'Resp':18,'WBC':8,
    'Creatinine':1,'Bilirubin_total':0.8,'Lactate':1.5,'Glucose':110,'Hgb':13,'pH':7.4,
    'Age':55,'HospAdmTime':-24,'ICULOS':2}
PRESETS={
    "🟢 Normal Patient":{'HR':78,'O2Sat':98,'Temp':36.8,'SBP':118,'MAP':82,'DBP':68,'Resp':15,
        'WBC':7.2,'Creatinine':0.9,'Bilirubin_total':0.7,'Lactate':1.1,'Glucose':95,'Hgb':13.5,
        'pH':7.41,'Age':45,'HospAdmTime':-24,'ICULOS':1},
    "🟡 Moderate Risk":{'HR':105,'O2Sat':92,'Temp':38.4,'SBP':94,'MAP':66,'DBP':56,'Resp':26,
        'WBC':15.8,'Creatinine':2.1,'Bilirubin_total':1.8,'Lactate':3.2,'Glucose':158,'Hgb':9.8,
        'pH':7.31,'Age':64,'HospAdmTime':-10,'ICULOS':4},
    "🔴 High Risk / Sepsis":{'HR':132,'O2Sat':86,'Temp':39.6,'SBP':74,'MAP':48,'DBP':38,'Resp':34,
        'WBC':24.1,'Creatinine':3.8,'Bilirubin_total':4.2,'Lactate':6.5,'Glucose':225,'Hgb':7.8,
        'pH':7.19,'Age':73,'HospAdmTime':-5,'ICULOS':2},
}
GROUPS=[("📊 Vital Signs",['HR','O2Sat','Temp','Resp']),("🩸 Blood Pressure",['SBP','DBP','MAP']),
    ("🔬 Laboratory",['WBC','Creatinine','Bilirubin_total','Lactate','Glucose','Hgb','pH']),
    ("👤 Patient Info",['Age','HospAdmTime','ICULOS'])]
ORGAN_FEATURES={'Heart':['HR','SBP','MAP','DBP'],'Lungs':['O2Sat','Resp','pH'],
    'Liver':['Bilirubin_total','WBC'],'Kidneys':['Creatinine','Glucose'],
    'Brain':['MAP','pH','Glucose'],'Blood':['WBC','Hgb','Lactate','Temp']}

# ═══ HELPERS ═══
def vital_status(feat,val):
    if feat not in CLINICAL_RANGES: return 'normal'
    mn,mx,_,_,_=CLINICAL_RANGES[feat]
    if feat=='O2Sat': return 'critical' if val<88 else 'abnormal' if val<95 else 'normal'
    if feat=='pH': return 'critical' if val<7.20 or val>7.55 else 'abnormal' if val<7.35 or val>7.45 else 'normal'
    if feat=='Lactate': return 'critical' if val>4 else 'abnormal' if val>2 else 'normal'
    if feat=='MAP': return 'critical' if val<50 else 'abnormal' if val<70 else 'normal'
    span=mx-mn
    if val<mn-span*0.5 or val>mx+span*0.5: return 'critical'
    if val<mn or val>mx: return 'abnormal'
    return 'normal'

def vital_color(s): return {'normal':'#059669','abnormal':'#d97706','critical':'#dc2626'}[s]
def vital_chip_class(s): return {'normal':'vc-normal','abnormal':'vc-warn','critical':'vc-critical'}[s]

def compute_sirs(inp):
    return {'Temp >38.3 or <36°C':inp.get('Temp',37)>38.3 or inp.get('Temp',37)<36,
            'HR > 90 bpm':inp.get('HR',80)>90,'Resp > 20 /min':inp.get('Resp',16)>20,
            'WBC >12 or <4 K/µL':inp.get('WBC',8)>12 or inp.get('WBC',8)<4}

def compute_sofa(inp):
    s=0; o2=inp.get('O2Sat',97)
    if o2<90: s+=3
    elif o2<94: s+=2
    elif o2<96: s+=1
    b=inp.get('Bilirubin_total',0.8)
    if b>=6: s+=3
    elif b>=2: s+=2
    elif b>=1.2: s+=1
    m=inp.get('MAP',85)
    if m<50: s+=3
    elif m<65: s+=2
    elif m<70: s+=1
    c=inp.get('Creatinine',1)
    if c>=3.5: s+=3
    elif c>=2: s+=2
    elif c>=1.2: s+=1
    return s

def shock_index(inp):
    hr=inp.get('HR',80);sbp=inp.get('SBP',120)
    return round(hr/sbp,2) if sbp>0 else 0.0

def clinical_recs(rp,inp,sm,sofa):
    recs=[]
    if rp>=60:
        recs.append(("high","⚠️ IMMEDIATE — SEP-1 PROTOCOL","Activate Sepsis Alert. Start 30 mL/kg IV crystalloid within 3 hours. Notify attending NOW."))
        recs.append(("high","🧫 CULTURES + LABS","Blood cultures ×2 BEFORE antibiotics. STAT: CBC, BMP, LFTs, coagulation, procalcitonin."))
        recs.append(("high","💊 ANTIBIOTICS",f"Broad-spectrum IV within 1 hour. MAP {inp.get('MAP',0):.0f} mmHg — target ≥65."))
        recs.append(("info","📊 MONITORING","Continuous vitals, foley catheter. Target UO ≥0.5 mL/kg/hr. Reassess q15min."))
    elif rp>=30:
        recs.append(("mod","⚡ INCREASED VIGILANCE",f"SIRS criteria met: {sm}/4. Reassess every 30 minutes."))
        recs.append(("mod","🔬 DIAGNOSTIC WORKUP","Order CBC, CMP, Lactate, Blood Cultures if febrile."))
        recs.append(("info","💧 FLUID ASSESSMENT","Assess volume responsiveness. IV fluid challenge 500 mL NS if hypotensive."))
    else:
        recs.append(("low","✅ ROUTINE MONITORING","Continue standard ICU monitoring. Reassess every 4 hours."))
        if sofa>=2: recs.append(("info","📋 SOFA NOTE",f"SOFA score {sofa} — document organ function baseline."))
    return recs

def organ_risk_pct(organ,inp,base_risk):
    feats=ORGAN_FEATURES.get(organ,[]);d=0
    for f in feats:
        if f in inp:
            s=vital_status(f,inp[f])
            if s=='critical': d+=2
            elif s=='abnormal': d+=1
    mx=len(feats)*2; bonus=(d/mx*0.25) if mx>0 else 0
    return min(100,round((base_risk+bonus)*100,1))

def risk_color_hex(pct):
    if pct<30: return f"#{int(5+(pct/30)*200):02x}{int(150-(pct/30)*20):02x}73"
    elif pct<60: t=(pct-30)/30; return f"#{int(205+t*50):02x}{int(130-t*80):02x}10"
    else: t=(pct-60)/40; return f"#{220:02x}{int(50-t*40):02x}0a"

BG_L='#ffffff';AX_L='#f8fafc';GRID_L=(0,0,0,0.06);TICK_L='#6a8090';TEXT_L='#1a2636'
def style_ax(ax,fig):
    fig.patch.set_facecolor(BG_L);ax.set_facecolor(AX_L)
    for s in ax.spines.values(): s.set_edgecolor((0,0,0,0.1))
    ax.spines['top'].set_visible(False);ax.spines['right'].set_visible(False)
    ax.tick_params(colors=TICK_L,labelsize=8.5);ax.xaxis.label.set_color(TICK_L);ax.yaxis.label.set_color(TICK_L)

# ═══ EXCEL EXPORT ═══
def generate_excel(inp,risk_pct,feat_imp,sirs_dict,sofa,si,organ_data, patient_name):
    buf=io.BytesIO()
    with pd.ExcelWriter(buf,engine='openpyxl') as w:
        pd.DataFrame([{'Feature':FEATURE_LABELS.get(k,k),'Value':v,
            'Unit':CLINICAL_RANGES.get(k,('','','','',''))[4],
            'Normal Min':CLINICAL_RANGES.get(k,(0,0,0,0,''))[0],
            'Normal Max':CLINICAL_RANGES.get(k,(0,0,0,0,''))[1],
            'Status':vital_status(k,v).upper()} for k,v in inp.items()
        ]).to_excel(w,sheet_name='Patient Vitals',index=False, startrow=3)
        pd.DataFrame([{'Metric':'Patient Name','Value':patient_name},
            {'Metric':'Sepsis Risk %','Value':f"{risk_pct}%"},
            {'Metric':'Risk Level','Value':'HIGH' if risk_pct>=60 else 'MODERATE' if risk_pct>=30 else 'LOW'},
            {'Metric':'SIRS Met','Value':f"{sum(sirs_dict.values())} / 4"},
            {'Metric':'SOFA Score','Value':sofa},{'Metric':'Shock Index','Value':si},
            {'Metric':'Timestamp','Value':datetime.now().strftime('%Y-%m-%d %H:%M')}
        ]).to_excel(w,sheet_name='Risk Summary',index=False, startrow=3)
        pd.DataFrame([{'Feature':FEATURE_LABELS.get(k,k),'SHAP Value':v,
            'Direction':'↑ Risk' if v>0 else '↓ Risk'}
            for k,v in sorted(feat_imp.items(),key=lambda x:abs(x[1]),reverse=True)
        ]).to_excel(w,sheet_name='SHAP Analysis',index=False, startrow=3)
        pd.DataFrame([{'Organ':o,'Risk %':f"{d['risk']}%",'Biomarkers':', '.join(ORGAN_FEATURES.get(o,[]))}
            for o,d in organ_data.items()]).to_excel(w,sheet_name='Organ Risk',index=False, startrow=3)
        pd.DataFrame([{'Criterion':k,'Met':'YES' if v else 'NO'} for k,v in sirs_dict.items()
        ]).to_excel(w,sheet_name='SIRS',index=False, startrow=3)
        
        # Apply professional styling to all sheets
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        workbook = w.book
        title_font = Font(bold=True, size=16, color='0C1F3A')
        subtitle_font = Font(italic=True, size=11, color='4A6080')
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0891B2', end_color='0891B2', fill_type='solid')
        thin_border = Border(left=Side(style='thin', color='DDE4ED'), right=Side(style='thin', color='DDE4ED'), top=Side(style='thin', color='DDE4ED'), bottom=Side(style='thin', color='DDE4ED'))
        
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            
            # Branding Headers
            ws['A1'] = f"Sepsis Early Warning System (EWS) - {sheet_name}"
            ws['A1'].font = title_font
            ws.merge_cells('A1:D1')
            
            ws['A2'] = f"Patient: {patient_name} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Auto-generated Report"
            ws['A2'].font = subtitle_font
            ws.merge_cells('A2:D2')
            
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.row == 4: # Table Header
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else: # Table Body
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = thin_border
                    
            for i in range(1, ws.max_column + 1):
                col_letter = get_column_letter(i)
                max_length = 0
                for cell in ws[col_letter]:
                    try:
                        if cell.value is not None:
                            cell_str = str(cell.value)
                            if '\n' in cell_str:
                                cell_str = max(cell_str.split('\n'), key=len)
                            max_length = max(max_length, len(cell_str))
                    except: pass
                ws.column_dimensions[col_letter].width = min(60, max_length + 2)
    buf.seek(0); return buf.read()

# ═══ PDF EXPORT ═══
def generate_pdf(inp,risk_pct,feat_imp,sirs_dict,sofa,si,organ_data, patient_name):
    buf=io.BytesIO()
    doc=SimpleDocTemplate(buf,pagesize=A4,topMargin=2*cm,bottomMargin=2*cm,leftMargin=2*cm,rightMargin=2*cm)
    styles=getSampleStyleSheet(); story=[]
    NAVY=rl_colors.HexColor('#0c1f3a');TEAL=rl_colors.HexColor('#0891b2')
    RED=rl_colors.HexColor('#dc2626');AMBER=rl_colors.HexColor('#d97706')
    GREEN=rl_colors.HexColor('#059669');LGRAY=rl_colors.HexColor('#f5f7fa');GRAY=rl_colors.HexColor('#dde4ed')
    rc=RED if risk_pct>=60 else AMBER if risk_pct>=30 else GREEN
    lt='HIGH RISK' if risk_pct>=60 else 'MODERATE RISK' if risk_pct>=30 else 'LOW RISK'
    ts=ParagraphStyle('T',fontName='Helvetica-Bold',fontSize=22,textColor=NAVY,spaceAfter=12,alignment=TA_CENTER)
    ss=ParagraphStyle('S',fontName='Helvetica',fontSize=10,textColor=rl_colors.HexColor('#4a6080'),alignment=TA_CENTER,spaceAfter=24)
    hs=ParagraphStyle('H',fontName='Helvetica-Bold',fontSize=12,textColor=TEAL,spaceBefore=16,spaceAfter=6)
    bs=ParagraphStyle('B',fontName='Helvetica',fontSize=10,textColor=rl_colors.HexColor('#1a2636'),leading=16,spaceAfter=6)
    story.append(Paragraph("SEPSIS EARLY WARNING SYSTEM",ts))
    story.append(Paragraph("Clinical Risk Assessment Report",ss))
    story.append(HRFlowable(width="100%",thickness=2,color=TEAL,spaceAfter=14))
    # Fix Patient Name Layout using Table
    local_time = datetime.now() + timedelta(hours=5)
    header_data = [[Paragraph(f"<b>Patient Name:</b> {patient_name}", bs),
                    Paragraph(f"<b>Date:</b> {local_time.strftime('%B %d, %Y %H:%M')}", ParagraphStyle('R', parent=bs, alignment=2))]] # alignment=2 is RIGHT
    header_table = Table(header_data, colWidths=[10*cm, 7*cm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10)
    ]))
    story.append(header_table)
    story.append(Spacer(1,10))
    rt=Table([['SEPSIS RISK',f'{risk_pct}%',lt],['SIRS',f"{sum(sirs_dict.values())}/4",'SOFA'],
        ['Shock Index',f'{si}',f'{sofa}']],colWidths=[5.5*cm]*3)
    rt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),NAVY),('TEXTCOLOR',(0,0),(-1,0),rl_colors.white),
        ('TEXTCOLOR',(1,0),(1,0),rc),('FONTNAME',(0,0),(-1,-1),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,0),14),('FONTSIZE',(0,1),(-1,-1),11),('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ROWBACKGROUNDS',(0,1),(-1,-1),[LGRAY,rl_colors.white]),
        ('GRID',(0,0),(-1,-1),.5,GRAY),('TOPPADDING',(0,0),(-1,-1),8),('BOTTOMPADDING',(0,0),(-1,-1),8)]))
    story.append(rt);story.append(Spacer(1,14))
    story.append(Paragraph("Patient Vital Signs",hs))
    vd=[['Feature','Value','Unit','Status']]
    for k,v in inp.items():
        vd.append([FEATURE_LABELS.get(k,k),f'{v:.2f}',CLINICAL_RANGES.get(k,('','','','',''))[4],vital_status(k,v).upper()])
    vt=Table(vd,colWidths=[5*cm,3.5*cm,3*cm,4*cm])
    vt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),TEAL),('TEXTCOLOR',(0,0),(-1,0),rl_colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTNAME',(0,1),(-1,-1),'Helvetica'),
        ('FONTSIZE',(0,0),(-1,-1),9),('ALIGN',(1,0),(-1,-1),'CENTER'),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[LGRAY,rl_colors.white]),('GRID',(0,0),(-1,-1),.4,GRAY),
        ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5)]))
    story.append(vt)
    story.append(Paragraph("SHAP Top 10",hs))
    ss2=sorted(feat_imp.items(),key=lambda x:abs(x[1]),reverse=True)[:10]
    sd=[['Feature','SHAP','Direction']]
    for k,v in ss2: sd.append([FEATURE_LABELS.get(k,k),f'{v:+.4f}','↑ Risk' if v>0 else '↓ Risk'])
    st2=Table(sd,colWidths=[5.5*cm,4*cm,6*cm])
    st2.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),TEAL),('TEXTCOLOR',(0,0),(-1,0),rl_colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTNAME',(0,1),(-1,-1),'Helvetica'),
        ('FONTSIZE',(0,0),(-1,-1),9),('ALIGN',(1,0),(-1,-1),'CENTER'),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[LGRAY,rl_colors.white]),('GRID',(0,0),(-1,-1),.4,GRAY),
        ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5)]))
    story.append(st2)
    story.append(Paragraph("Organ Risk",hs))
    od2=[['Organ','Risk %','Status','Biomarkers']]
    for o,d in organ_data.items():
        r=d['risk'];s='CRITICAL' if r>=60 else 'ELEVATED' if r>=30 else 'NORMAL'
        od2.append([o,f'{r}%',s,', '.join(ORGAN_FEATURES.get(o,[]))])
    ot=Table(od2,colWidths=[3.5*cm,2.5*cm,3.5*cm,6*cm])
    ot.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),TEAL),('TEXTCOLOR',(0,0),(-1,0),rl_colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTNAME',(0,1),(-1,-1),'Helvetica'),
        ('FONTSIZE',(0,0),(-1,-1),9),('ALIGN',(1,0),(2,-1),'CENTER'),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[LGRAY,rl_colors.white]),('GRID',(0,0),(-1,-1),.4,GRAY),
        ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5)]))
    story.append(ot);story.append(Spacer(1,18));story.append(HRFlowable(width="100%",thickness=1,color=GRAY))
    story.append(Paragraph("Sepsis Early Warning System · Confidential Clinical Record",
        ParagraphStyle('f',fontName='Helvetica',fontSize=8,textColor=rl_colors.HexColor('#8aa0b8'),alignment=TA_CENTER,spaceBefore=8)))
    doc.build(story);buf.seek(0);return buf.read()

# ═══ REALISTIC ANATOMICAL BODY MAP (SVG) ═══
def render_body_map(organ_data, selected_organ, height=580):
    def risk_color(o):
        r = organ_data[o]['risk']
        return '#dc2626' if r >= 60 else '#d97706' if r >= 30 else '#059669'
    
    def line_style(o):
        is_sel = (o == selected_organ)
        color = '#ffffff' if is_sel else 'rgba(100, 116, 139, 0.4)'
        width = '4' if is_sel else '2'
        shadow = 'filter="url(#glow)"' if is_sel else ''
        anim = 'style="animation: pulseOrgan 2.5s infinite;"' if is_sel else ''
        return f'stroke="{color}" stroke-width="{width}" {shadow} {anim}'

    def text_style(o):
        is_sel = (o == selected_organ)
        color = '#ffffff' if is_sel else '#64748b'
        weight = '900' if is_sel else 'bold'
        size = '28' if is_sel else '22'
        shadow = 'filter="url(#glow)"' if is_sel else ''
        anim = 'style="animation: pulseOrgan 2.5s infinite;"' if is_sel else ''
        return f'fill="{color}" font-weight="{weight}" font-size="{size}" {shadow} {anim}'

    import base64
    import os
    img_path = os.path.join(os.path.dirname(__file__), "latest_stitch_body.png")
    bg_style = 'background:radial-gradient(circle at 50% 50%, #1e293b 0%, #020617 100%);border-radius:16px;border:1px solid #334155;box-shadow:0 10px 30px rgba(0,0,0,0.5)'
    if os.path.exists(img_path):
        with open(img_path, "rb") as f:
            b64_img = base64.b64encode(f.read()).decode("utf-8")
        bg_style = f'background: url(data:image/png;base64,{b64_img}) no-repeat center center; background-size: cover; border-radius:16px; border:1px solid #334155; box-shadow:0 10px 30px rgba(0,0,0,0.5)'

    svg = f"""<!DOCTYPE html><html><head><style>
    @keyframes pulseOrgan {{ 0% {{ filter: drop-shadow(0 0 10px rgba(255,255,255,0.8)); opacity: 1; }} 50% {{ filter: drop-shadow(0 0 25px rgba(255,255,255,1)); opacity: 1; }} 100% {{ filter: drop-shadow(0 0 10px rgba(255,255,255,0.8)); opacity: 1; }} }}
    .organ {{ transition: all 0.3s ease; }}
    </style></head><body style="margin:0;padding:0;background:transparent;display:flex;justify-content:center;align-items:center;">
<svg viewBox="0 0 768 1376" xmlns="http://www.w3.org/2000/svg" width="100%" style="max-width:320px;height:auto;display:block;margin:auto;{bg_style}">
<defs>
  <filter id="glow" x="-20%" y="-20%" width="140%" height="140%"><feGaussianBlur stdDeviation="8" result="blur"/><feComposite in="SourceGraphic" in2="blur" operator="over"/></filter>
</defs>

<!-- LEADER LABELS & CONNECTORS -->
<g font-family="'JetBrains Mono', monospace">
  <!-- BRAIN -->
  <text x="140" y="128" text-anchor="end" {text_style('Brain')}>BRAIN</text>
  <line x1="150" y1="120" x2="384" y2="120" {line_style('Brain')} stroke-dasharray="4,6" stroke-linecap="round"/>
  <rect x="200" y="100" width="70" height="40" rx="8" fill="#0f172a" stroke="{risk_color('Brain') if selected_organ == 'Brain' else '#475569'}" stroke-width="{ '4' if selected_organ == 'Brain' else '2' }"/>
  <text x="235" y="128" text-anchor="middle" font-size="22" font-weight="bold" fill="{risk_color('Brain')}">{organ_data['Brain']['risk']}%</text>

  <!-- LUNGS -->
  <text x="140" y="383" text-anchor="end" {text_style('Lungs')}>LUNGS</text>
  <line x1="150" y1="375" x2="340" y2="375" {line_style('Lungs')} stroke-dasharray="4,6" stroke-linecap="round"/>
  <rect x="180" y="355" width="70" height="40" rx="8" fill="#0f172a" stroke="{risk_color('Lungs') if selected_organ == 'Lungs' else '#475569'}" stroke-width="{ '4' if selected_organ == 'Lungs' else '2' }"/>
  <text x="215" y="383" text-anchor="middle" font-size="22" font-weight="bold" fill="{risk_color('Lungs')}">{organ_data['Lungs']['risk']}%</text>

  <!-- LIVER -->
  <text x="140" y="516" text-anchor="end" {text_style('Liver')}>LIVER</text>
  <line x1="150" y1="508" x2="322" y2="508" {line_style('Liver')} stroke-dasharray="4,6" stroke-linecap="round"/>
  <rect x="180" y="488" width="70" height="40" rx="8" fill="#0f172a" stroke="{risk_color('Liver') if selected_organ == 'Liver' else '#475569'}" stroke-width="{ '4' if selected_organ == 'Liver' else '2' }"/>
  <text x="215" y="516" text-anchor="middle" font-size="22" font-weight="bold" fill="{risk_color('Liver')}">{organ_data['Liver']['risk']}%</text>

  <!-- HEART -->
  <text x="628" y="423" text-anchor="start" {text_style('Heart')}>HEART</text>
  <line x1="618" y1="415" x2="416" y2="415" {line_style('Heart')} stroke-dasharray="4,6" stroke-linecap="round"/>
  <rect x="498" y="395" width="70" height="40" rx="8" fill="#0f172a" stroke="{risk_color('Heart') if selected_organ == 'Heart' else '#475569'}" stroke-width="{ '4' if selected_organ == 'Heart' else '2' }"/>
  <text x="533" y="423" text-anchor="middle" font-size="22" font-weight="bold" fill="{risk_color('Heart')}">{organ_data['Heart']['risk']}%</text>

  <!-- KIDNEYS -->
  <text x="628" y="572" text-anchor="start" {text_style('Kidneys')}>KIDNEYS</text>
  <line x1="618" y1="564" x2="430" y2="564" {line_style('Kidneys')} stroke-dasharray="4,6" stroke-linecap="round"/>
  <rect x="498" y="544" width="70" height="40" rx="8" fill="#0f172a" stroke="{risk_color('Kidneys') if selected_organ == 'Kidneys' else '#475569'}" stroke-width="{ '4' if selected_organ == 'Kidneys' else '2' }"/>
  <text x="533" y="572" text-anchor="middle" font-size="22" font-weight="bold" fill="{risk_color('Kidneys')}">{organ_data['Kidneys']['risk']}%</text>

  <!-- BLOOD -->
  <text x="628" y="808" text-anchor="start" {text_style('Blood')}>BLOOD</text>
  <line x1="618" y1="800" x2="430" y2="800" {line_style('Blood')} stroke-dasharray="4,6" stroke-linecap="round"/>
  <rect x="540" y="780" width="70" height="40" rx="8" fill="#0f172a" stroke="{risk_color('Blood') if selected_organ == 'Blood' else '#475569'}" stroke-width="{ '4' if selected_organ == 'Blood' else '2' }"/>
  <text x="575" y="808" text-anchor="middle" font-size="22" font-weight="bold" fill="{risk_color('Blood')}">{organ_data['Blood']['risk']}%</text>
</g>
</svg></body></html>"""
    return svg


with st.sidebar:
    st.markdown("""<div class='sidebar-header'>
        <div class='sidebar-icon'>🏥</div>
        <div class='sidebar-title'>PATIENT INPUT</div>
        <div class='sidebar-sub'>ICU Clinical Parameters</div>
    </div>""", unsafe_allow_html=True)
    patient_name = st.text_input("Patient Name", value=st.session_state.get('patient_name', 'John Doe'))
    st.session_state.patient_name = patient_name
    input_mode=st.radio("Input method",["Manual entry","Load preset scenario","Upload .txt file"],horizontal=False)
    inputs={}
    if input_mode=="Load preset scenario":
        preset=st.selectbox("Scenario",list(PRESETS.keys()));inputs={k:float(v) for k,v in PRESETS[preset].items()}
        for f in features:
            if f not in inputs: inputs[f]=float(DEFAULTS.get(f,0))
        st.success(f"✅ {len(inputs)} values loaded")
    elif input_mode=="Upload .txt file":
        st.markdown("<div style='font-size:11px;color:var(--text2);margin-bottom:8px'>Format: <code>FeatureName: Value</code> per line</div>",unsafe_allow_html=True)
        uploaded=st.file_uploader("Upload .txt",type=["txt"])
        if uploaded:
            lines=uploaded.read().decode().strip().split('\n')
            parsed=0
            for line in lines:
                if ':' in line or '=' in line:
                    sep=':' if ':' in line else '='
                    k,v=line.split(sep,1)
                    k,v=k.strip(),v.strip()
                    if k in features:
                        try:
                            inputs[k]=float(v); parsed+=1
                        except: pass
            for f in features:
                if f not in inputs: inputs[f]=float(DEFAULTS.get(f,0))
            st.success(f"✅ Loaded {parsed} valid values from file")
            st.info(f"ℹ️ {len(features)-parsed} missing features filled with defaults")
        else:
            example_txt = "\n".join([f"{f}: {DEFAULTS.get(f,0)}" for f in features])
            st.download_button("📥 Download Example Format", data=example_txt, file_name="sepsis_patient_example.txt", mime="text/plain", use_container_width=True)
            for f in features: inputs[f]=float(DEFAULTS.get(f,0))
    else:
        for grp,feats in GROUPS:
            with st.expander(grp,expanded=(grp=="📊 Vital Signs")):
                for feat in feats:
                    lbl=FEATURE_LABELS.get(feat,feat);cfg=CLINICAL_RANGES.get(feat,(0,200,0,300,''));unit=cfg[4]
                    disp=f"{lbl} [{unit}]" if unit else lbl
                    inputs[feat]=st.number_input(disp,min_value=float(cfg[2]),max_value=float(cfg[3]),
                        value=float(DEFAULTS.get(feat,(cfg[2]+cfg[3])/2)),step=0.1,key=f"inp_{feat}")
    st.markdown("<br>",unsafe_allow_html=True)
    predict_btn=st.button("⚡  Analyze Sepsis Risk",type="primary",use_container_width=True)
    st.markdown("""<hr style='border:1px solid #dde4ed;margin:14px 0'>
    <div style='font-size:10.5px;color:#8aa0b8;line-height:2;font-family:JetBrains Mono'>
        <b style='color:#4a6080'>Risk Thresholds</b><br>🟢 &lt;30% Low · 🟡 30–60% Moderate · 🔴 &gt;60% High<br></div>""",unsafe_allow_html=True)


# ═══ HEADER ═══
st.markdown(f"""<div class="hdr-wrap">
    <div class="hdr-title">Sepsis Early Warning System</div>
    <div class="hdr-sub">ICU Clinical Decision Support — Real-Time Risk Stratification</div>
    <div><span class="hdr-badge hdr-live">● LIVE MONITORING</span></div>
</div>""", unsafe_allow_html=True)

# ═══ PREDICTION ═══
if 'risk_pct' not in st.session_state: st.session_state.risk_pct=None
if predict_btn:
    if 'pdf_data' in st.session_state: del st.session_state.pdf_data
    if 'xlsx_data' in st.session_state: del st.session_state.xlsx_data
    arr=np.array([[inputs.get(f,0) for f in features]])
    arr_s=scaler.transform(arr);raw_prob=float(model.predict_proba(arr_s)[0][1])
    # Calibrate probability to counteract balanced training dataset squashing
    prob = float(np.interp(raw_prob, [0.0, 0.30, 0.68, 0.70, 1.0], [0.0, 0.05, 0.45, 0.85, 1.0]))
    st.session_state.risk_pct=round(prob*100,1)
    shap_vals=explainer.shap_values(arr_s)
    sv=shap_vals[0] if isinstance(shap_vals,list) else shap_vals[0]
    st.session_state.shap_dict={features[i]:float(sv[i]) for i in range(len(features))}
    st.session_state.inputs=dict(inputs)

if st.session_state.risk_pct is not None:
    rp=st.session_state.risk_pct;inp=st.session_state.inputs;feat_imp=st.session_state.shap_dict
    sirs_dict=compute_sirs(inp);sirs_met=sum(sirs_dict.values());sofa=compute_sofa(inp);si=shock_index(inp)
    risk_cls='risk-high' if rp>=60 else 'risk-mod' if rp>=30 else 'risk-low'
    risk_lbl='HIGH RISK' if rp>=60 else 'MODERATE' if rp>=30 else 'LOW RISK'
    risk_msg='Immediate clinical intervention required. Activate SEP-1.' if rp>=60 else 'Close monitoring recommended. Reassess in 30 min.' if rp>=30 else 'Continue routine monitoring every 4 hours.'
    prog_clr='#dc2626' if rp>=60 else '#d97706' if rp>=30 else '#059669'

    organ_data={o:{'risk':organ_risk_pct(o,inp,rp/100)} for o in ORGAN_FEATURES}

    # ── Stats row
    c1,c2,c3,c4,c5=st.columns(5)
    local_time_str = (datetime.now() + timedelta(hours=5)).strftime('%H:%M:%S')
    stats=[('Sepsis Risk',f'{rp}%','risk-stat'),('SIRS Score',f'{sirs_met}/4','sirs-stat'),
           ('SOFA Score',f'{sofa}','sofa-stat'),('Shock Index',f'{si}','si-stat'),
           ('Timestamp',local_time_str,'time-stat')]
    for col,(lbl,val,_) in zip([c1,c2,c3,c4,c5],stats):
        col.markdown(f'<div class="stat-card"><div class="stat-num">{val}</div><div class="stat-lbl">{lbl}</div></div>',unsafe_allow_html=True)

    st.markdown(f"""<div class="prog-track"><div class="prog-fill" style="width:{min(rp,100)}%;background:{prog_clr}"></div></div>
    <div class="prog-labels"><span>0% Safe</span><span>30% Moderate</span><span>60% High</span><span>100% Critical</span></div>""",unsafe_allow_html=True)
    st.markdown("")

    # ── Tabs
    tabs=st.tabs(["🎯 Risk Overview","📊 SHAP Analysis","🧍🏻 Body Map","📋 Clinical Scores","📥 Export"])

    # ──── TAB 0: Risk Overview ────
    with tabs[0]:
        col_r,col_v=st.columns([1,2])
        with col_r:
            st.markdown(f"""<div class="risk-card {risk_cls}">
                <div class="risk-pct">{rp}%</div><div class="risk-level">{risk_lbl}</div>
                <div class="risk-msg">{risk_msg}</div></div>""",unsafe_allow_html=True)
            st.markdown("<br>",unsafe_allow_html=True)
            recs=clinical_recs(rp,inp,sirs_met,sofa)
            for sev,title,msg in recs:
                cls={'high':'rc-high','mod':'rc-mod','low':'rc-low','info':'rc-info'}[sev]
                st.markdown(f'<div class="rec-card {cls}"><div class="rec-title">{title}</div>{msg}</div>',unsafe_allow_html=True)
        with col_v:
            st.markdown('<div class="s-hdr">Patient Vital Signs</div>',unsafe_allow_html=True)
            vital_feats=[f for f in inp if f in CLINICAL_RANGES and f not in ('Age','HospAdmTime','ICULOS')]
            cols_per_row=4
            for i in range(0,len(vital_feats),cols_per_row):
                batch=vital_feats[i:i+cols_per_row]
                cols=st.columns(cols_per_row)
                for j,feat in enumerate(batch):
                    v=inp[feat];s=vital_status(feat,v);cfg=CLINICAL_RANGES[feat]
                    chip_cls=vital_chip_class(s);vc=vital_color(s)
                    lbl=FEATURE_LABELS.get(feat,feat);unit=cfg[4]
                    cols[j].markdown(f"""<div class="vital-chip {chip_cls}">
                        <div class="vital-name">{lbl}</div>
                        <div class="vital-val" style="color:{vc}">{v:.1f}</div>
                        <div class="vital-unit">{unit}</div>
                        <div class="vital-range">N: {cfg[0]}-{cfg[1]}</div></div>""",unsafe_allow_html=True)
            st.markdown("<br>",unsafe_allow_html=True)
            st.markdown('<div class="s-hdr">SIRS Criteria</div>',unsafe_allow_html=True)
            sirs_html='<div class="sirs-wrap">'
            for crit,met in sirs_dict.items():
                dot_c='#059669' if met else '#dc2626'
                sirs_html+=f'<div class="sirs-row"><div class="sirs-dot" style="background:{dot_c}"></div><span>{crit}</span><span style="margin-left:auto;font-weight:700;color:{dot_c}">{"MET" if met else "NOT MET"}</span></div>'
            sirs_html+='</div>'
            st.markdown(sirs_html,unsafe_allow_html=True)
            sts='a-high' if sirs_met>=3 else 'a-mod' if sirs_met>=2 else 'a-low'
            st.markdown(f'<div class="alert-strip {sts}">SIRS: {sirs_met}/4 criteria met</div>',unsafe_allow_html=True)


    # ──── TAB 1: SHAP Analysis ────
    with tabs[1]:
        st.markdown('<div class="s-hdr">SHAP Feature Importance — XGBoost Explainability</div>',unsafe_allow_html=True)
        sorted_feats=sorted(feat_imp.items(),key=lambda x:abs(x[1]),reverse=True)
        top_n=min(12,len(sorted_feats))
        top=sorted_feats[:top_n]; labels=[FEATURE_LABELS.get(k,k) for k,_ in reversed(top)]
        vals=[v for _,v in reversed(top)]

        fig,ax=plt.subplots(figsize=(10,5)); style_ax(ax,fig)
        colors=['#dc2626' if v>0 else '#059669' for v in vals]
        bars=ax.barh(range(len(labels)),vals,color=colors,height=0.6,edgecolor='none',alpha=0.85)
        ax.set_yticks(range(len(labels)));ax.set_yticklabels(labels,fontsize=9,fontfamily='monospace')
        ax.set_xlabel('SHAP Value (Impact on Prediction)',fontsize=10)
        ax.set_title('Feature Contributions to Sepsis Risk',fontsize=13,fontweight='bold',color='#1a2636',pad=14)
        ax.axvline(x=0,color='#8aa0b8',linewidth=0.8,alpha=0.4)
        ax.grid(axis='x',alpha=0.08);ax.tick_params(axis='y',length=0)
        red_patch=mpatches.Patch(color='#dc2626',label='↑ Increases Risk')
        green_patch=mpatches.Patch(color='#059669',label='↓ Decreases Risk')
        ax.legend(handles=[red_patch,green_patch],loc='lower right',fontsize=8,framealpha=0.8)
        fig.tight_layout(); st.pyplot(fig); plt.close(fig)

        st.markdown("**Top 5 Risk Drivers:**")
        for k,v in sorted_feats[:5]:
            dr="↑ increases" if v>0 else "↓ decreases"
            st.markdown(f"- **{FEATURE_LABELS.get(k,k)}** ({v:+.4f}) — {dr} sepsis risk")

    # ──── TAB 2: Body Map ────
    with tabs[2]:
        st.markdown('<div class="s-hdr">Anatomical Risk Visualization</div>',unsafe_allow_html=True)
        st.markdown('<div class="i-box">Select an organ below to inspect its risk level and associated biomarkers.</div>',unsafe_allow_html=True)
        st.markdown("")
        organ_names=list(ORGAN_FEATURES.keys())
        sel_organ=st.radio("Select Organ",organ_names,horizontal=True,key='organ_sel')

        col_map,col_det=st.columns([1,1])
        with col_map:
            svg_html=render_body_map(organ_data,sel_organ)
            st_html.html(svg_html, height=580, scrolling=False)
        with col_det:
            od=organ_data[sel_organ]; risk_v=od['risk']
            oc='#dc2626' if risk_v>=60 else '#d97706' if risk_v>=30 else '#059669'
            olbl='CRITICAL' if risk_v>=60 else 'ELEVATED' if risk_v>=30 else 'NORMAL'
            st.markdown(f"""<div class="organ-risk-card" style="border-color: {oc};">
                <div class="organ-risk-title">{sel_organ}</div>
                <div class="organ-risk-value" style="color: {oc};">{risk_v}%</div>
                <div class="organ-risk-badge" style="background: {oc}15; color: {oc}; border-color: {oc}44;">{olbl}</div>
            </div>""",unsafe_allow_html=True)
            st.markdown("<br>",unsafe_allow_html=True)
            st.markdown(f'<div class="s-hdr">{sel_organ} — Associated Biomarkers</div>',unsafe_allow_html=True)
            for bf in ORGAN_FEATURES.get(sel_organ,[]):
                if bf in inp:
                    bv=inp[bf];bs=vital_status(bf,bv);bc=vital_color(bs);cfg=CLINICAL_RANGES.get(bf,('','','','',''))
                    st.markdown(f"""<div class="biomarker-row" style="border-left-color: {bc};">
                        <span class="biomarker-label">{FEATURE_LABELS.get(bf,bf)}</span>
                        <span class="biomarker-val" style="color:{bc}">{bv:.1f} <span style="font-size:10px;opacity:0.7;font-weight:600">{cfg[4]}</span></span>
                    </div>""",unsafe_allow_html=True)

    # ──── TAB 3: Clinical Scores ────
    with tabs[3]:
        st.markdown('<div class="s-hdr">Clinical Scoring Systems</div>',unsafe_allow_html=True)
        cs1,cs2,cs3=st.columns(3)
        cs1.markdown(f"""<div class="score-box" style="border-top:3px solid {'#dc2626' if sofa>=6 else '#d97706' if sofa>=2 else '#059669'}">
            <div class="score-num" style="color:{'#dc2626' if sofa>=6 else '#d97706' if sofa>=2 else '#059669'}">{sofa}</div>
            <div class="score-lbl">SOFA Score</div></div>""",unsafe_allow_html=True)
        cs2.markdown(f"""<div class="score-box" style="border-top:3px solid {'#dc2626' if sirs_met>=3 else '#d97706' if sirs_met>=2 else '#059669'}">
            <div class="score-num" style="color:{'#dc2626' if sirs_met>=3 else '#d97706' if sirs_met>=2 else '#059669'}">{sirs_met}/4</div>
            <div class="score-lbl">SIRS Criteria</div></div>""",unsafe_allow_html=True)
        si_c='#dc2626' if si>1.0 else '#d97706' if si>0.7 else '#059669'
        cs3.markdown(f"""<div class="score-box" style="border-top:3px solid {si_c}">
            <div class="score-num" style="color:{si_c}">{si}</div>
            <div class="score-lbl">Shock Index</div></div>""",unsafe_allow_html=True)

        st.markdown("")
        st.markdown('<div class="s-hdr">Score Breakdown</div>',unsafe_allow_html=True)
        detail_html='<div class="scores-box">'
        score_items=[("O₂ Saturation",f"{inp.get('O2Sat',0):.0f}%"),
            ("Bilirubin",f"{inp.get('Bilirubin_total',0):.1f} mg/dL"),
            ("MAP",f"{inp.get('MAP',0):.0f} mmHg"),
            ("Creatinine",f"{inp.get('Creatinine',0):.1f} mg/dL"),
            ("Lactate",f"{inp.get('Lactate',0):.1f} mmol/L"),
            ("Heart Rate",f"{inp.get('HR',0):.0f} bpm"),
            ("WBC Count",f"{inp.get('WBC',0):.1f} K/µL")]
        for lbl,val in score_items:
            detail_html+=f'<div class="score-row"><span class="score-label">{lbl}</span><span class="score-value">{val}</span></div>'
        detail_html+='</div>'
        st.markdown(detail_html,unsafe_allow_html=True)

        st.markdown("")
        st.markdown('<div class="s-hdr">Sepsis-3 Assessment</div>',unsafe_allow_html=True)
        s3_met=sofa>=2 and (inp.get('MAP',85)<65 or inp.get('Lactate',1)>2)
        if s3_met:
            st.markdown('<div class="alert-strip a-high">⚠️ SEPSIS-3 CRITERIA MET — Suspected Sepsis with Organ Dysfunction</div>',unsafe_allow_html=True)
        elif sofa>=2:
            st.markdown('<div class="alert-strip a-mod">⚡ SOFA ≥ 2 — Monitor for Sepsis Progression</div>',unsafe_allow_html=True)
        else:
            st.markdown('<div class="alert-strip a-low">✅ Sepsis-3 Criteria NOT Met</div>',unsafe_allow_html=True)

    # ──── TAB 4: Export ────
    with tabs[4]:
        st.markdown('<div class="s-hdr">Clinical Report Export</div>',unsafe_allow_html=True)
        st.markdown('<div class="i-box">Download comprehensive clinical reports with all patient data, SHAP analysis, organ risk assessment, and scoring system details.</div>',unsafe_allow_html=True)
        st.markdown("")
        ex1,ex2=st.columns(2)
        with ex1:
            if st.button("🛠️ Generate PDF Report", use_container_width=True):
                with st.spinner("Generating PDF..."):
                    st.session_state.pdf_data = generate_pdf(inp,rp,feat_imp,sirs_dict,sofa,si,organ_data, st.session_state.patient_name)
            
            if "pdf_data" in st.session_state:
                st.download_button("📄 Download PDF",data=st.session_state.pdf_data,file_name=f"sepsis_report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                    mime="application/pdf",use_container_width=True,type="primary")
        with ex2:
            if st.button("🛠️ Generate Excel Report", use_container_width=True):
                with st.spinner("Generating Excel..."):
                    st.session_state.xlsx_data = generate_excel(inp,rp,feat_imp,sirs_dict,sofa,si,organ_data, st.session_state.patient_name)
            
            if "xlsx_data" in st.session_state:
                st.download_button("📊 Download Excel",data=st.session_state.xlsx_data,file_name=f"sepsis_data_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
        st.markdown("")
        st.markdown("**Report includes:** Patient vitals, risk assessment, SHAP analysis, organ risk mapping, SIRS/SOFA/Shock Index scores, and clinical recommendations.")

    # ── Summary bar
    sb_cls='a-high' if rp>=60 else 'a-mod' if rp>=30 else 'a-low'
    emoji='🔴' if rp>=60 else '🟡' if rp>=30 else '🟢'
    st.markdown(f'<div class="summary-bar {sb_cls}">{emoji}  SEPSIS RISK: {rp}%  •  SIRS: {sirs_met}/4  •  SOFA: {sofa}  •  SI: {si}  •  {risk_lbl}</div>',unsafe_allow_html=True)

else:
    # ── Default landing
    st.markdown("")
    how_cols=st.columns(4)
    steps=[("1","Enter Vitals","Input parameters via sidebar"),("2","Run Analysis","XGBoost + SHAP prediction"),
           ("3","Review Results","Risk scores, body map, vitals"),("4","Export Report","Download PDF or Excel")]
    for col,(n,t,d) in zip(how_cols,steps):
        col.markdown(f'<div class="step-card"><div class="step-n">{n}</div><div class="step-t">{t}</div><div class="step-d">{d}</div></div>',unsafe_allow_html=True)
    st.markdown('<div class="alert-strip a-info">ℹ️  Enter patient vitals and click "Analyze Sepsis Risk" to begin</div>',unsafe_allow_html=True)

# ── Footer
st.markdown(f"""<div class="footer">
    🏥 Sepsis Early Warning System v6.0 &nbsp;·&nbsp; Clinical Decision Support Tool<br>
    Confidential Patient Record &nbsp;·&nbsp; {(datetime.now() + timedelta(hours=5)).strftime('%B %d, %Y | %I:%M %p')}
</div>""", unsafe_allow_html=True)
